#!/usr/bin/env python3
"""
Generate the Bulk Create shipments template (.xlsx).

This is the static asset served by the dashboard
  Bulk Create → "Download template" button.

The "Shipments" tab mirrors the FETCHER B2B BOOKING sheet — only the
user-fillable columns, in the sheet's exact left-to-right order. Enum columns
are real Excel dropdowns; NO OF PIECES / WEIGHT have numeric validation. A
second "Instructions" tab documents how to fill it.

Auto / ops-controlled columns (DATE, AWB NO, BATCH NO, CLIENT ID,
CHARGEABLE WEIGHT, STATUS, REMARKS, ESTIMATED DELIVERY DATE, BILLING AMOUNT)
are intentionally excluded — they are assigned automatically or by ops.

Run:  python3 scripts/generate-bulk-template.py
Out:  public/templates/fetcher-bulk-shipments-template.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ---- Brand palette (from globals.css @theme) ----
PURPLE = "2B296A"  # required-column header
GRAY = "585858"    # optional-column header
ORANGE = "F08C2A"
LIGHT = "F2F4F6"
DARK = "141006"

MAX_ROWS = 500  # number of data rows the dropdowns / validation cover

OUT_PATH = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "public", "templates", "fetcher-bulk-shipments-template.xlsx",
)

# 5 parcel column-sets; only Parcel 1's pieces + weight are required.
_PARCEL_COLUMNS = []
for _n in range(1, 6):
    _PARCEL_COLUMNS += [
        (f"PARCEL {_n} NO OF PIECES", _n == 1, None),
        (f"PARCEL {_n} WEIGHT (IN KG)", _n == 1, None),
        (f"PARCEL {_n} DIMENSIONS (IN CM)", False, None),
    ]

# (header, required?, dropdown values or None)
COLUMNS = [
    ("REF", False, None),
    ("SCOPE", True, ["Domestic", "International"]),
    ("PICK UP ADDRESS", True, None),
    ("PICK UP PINCODE/ZIP CODE", True, None),
    ("PICK UP CONTACT PERSON", False, None),
    ("PICK UP CONTACT NO", True, None),
    ("PICK UP CONTACT EMAIL", False, None),
    ("PICK UP ALTERNATE CONTACT PERSON", False, None),
    ("PICK UP ALTERNATE CONTACT NO", False, None),
    *_PARCEL_COLUMNS,
    ("DELIVERY ADDRESS", True, None),
    ("DELIVERY PINCODE/ZIP CODE", True, None),
    ("DELIVERY CONTACT NO", True, None),
    ("DELIVERY CONTACT PERSON", False, None),
    ("DELIVERY CONTACT EMAIL", False, None),
    ("DELIVERY ALTERNATE CONTACT PERSON", False, None),
    ("DELIVERY ALTERNATE CONTACT NO", False, None),
    ("SHIPMENT TYPE", True, ["Commercial", "Non-Commercial"]),
    ("MODE", True, ["Express", "Air", "Surface", "Eco-Ground"]),
    ("SHIPMENT CATAGORY", True, ["Doc", "Non-Doc"]),  # sheet's spelling, kept to match
    ("IS THIS A DG SHIPMENT?", False, ["Yes", "No"]),
    ("ADDITIONAL INFORMATION", False, None),
]

WIDTHS = {
    "REF": 12,
    "SCOPE": 16,
    "PICK UP ADDRESS": 40,
    "PICK UP PINCODE/ZIP CODE": 20,
    "PICK UP CONTACT PERSON": 22,
    "PICK UP CONTACT NO": 18,
    "PICK UP CONTACT EMAIL": 26,
    "PICK UP ALTERNATE CONTACT PERSON": 28,
    "PICK UP ALTERNATE CONTACT NO": 26,
    "DELIVERY ADDRESS": 40,
    "DELIVERY PINCODE/ZIP CODE": 20,
    "DELIVERY CONTACT NO": 18,
    "DELIVERY CONTACT PERSON": 22,
    "DELIVERY CONTACT EMAIL": 26,
    "DELIVERY ALTERNATE CONTACT PERSON": 28,
    "DELIVERY ALTERNATE CONTACT NO": 26,
    "SHIPMENT TYPE": 18,
    "MODE": 16,
    "SHIPMENT CATAGORY": 18,
    "IS THIS A DG SHIPMENT?": 22,
    "ADDITIONAL INFORMATION": 34,
}
for _n in range(1, 6):
    WIDTHS[f"PARCEL {_n} NO OF PIECES"] = 16
    WIDTHS[f"PARCEL {_n} WEIGHT (IN KG)"] = 17
    WIDTHS[f"PARCEL {_n} DIMENSIONS (IN CM)"] = 18

NOTES = {
    "REF": "Your own reference (optional) — echoed back in the results so you can match errors to rows.",
    "SCOPE": "Dropdown: Domestic / International",
    "PICK UP ADDRESS": "Full pickup address",
    "PICK UP PINCODE/ZIP CODE": "Pickup pincode / ZIP code",
    "PICK UP CONTACT PERSON": "Contact name at pickup (optional)",
    "PICK UP CONTACT NO": "Contact phone at pickup",
    "PICK UP CONTACT EMAIL": "Contact email at pickup (optional)",
    "PICK UP ALTERNATE CONTACT PERSON": "Backup contact name at pickup (optional)",
    "PICK UP ALTERNATE CONTACT NO": "Backup contact phone at pickup (optional)",
    "DELIVERY ADDRESS": "Full delivery address",
    "DELIVERY PINCODE/ZIP CODE": "Delivery pincode / ZIP code",
    "DELIVERY CONTACT NO": "Contact phone at delivery",
    "DELIVERY CONTACT PERSON": "Contact name at delivery (optional)",
    "DELIVERY CONTACT EMAIL": "Contact email at delivery (optional)",
    "DELIVERY ALTERNATE CONTACT PERSON": "Backup contact name at delivery (optional)",
    "DELIVERY ALTERNATE CONTACT NO": "Backup contact phone at delivery (optional)",
    "SHIPMENT TYPE": "Dropdown: Commercial / Non-Commercial",
    "MODE": "Dropdown: Express / Air / Surface / Eco-Ground",
    "SHIPMENT CATAGORY": "Dropdown: Doc / Non-Doc",
    "IS THIS A DG SHIPMENT?": "Dropdown: Yes / No (blank = No)",
    "ADDITIONAL INFORMATION": "Any extra notes (optional)",
}
for _n in range(1, 6):
    _suffix = "" if _n == 1 else " — only if this parcel is used"
    NOTES[f"PARCEL {_n} NO OF PIECES"] = "Whole number, at least 1" + _suffix
    NOTES[f"PARCEL {_n} WEIGHT (IN KG)"] = "Number greater than 0" + _suffix
    NOTES[f"PARCEL {_n} DIMENSIONS (IN CM)"] = "e.g. 30x20x15 (optional)"


def build_shipments_sheet(ws):
    req_fill = PatternFill("solid", fgColor=PURPLE)
    opt_fill = PatternFill("solid", fgColor=GRAY)
    head_font = Font(bold=True, color="FFFFFF", size=11)
    head_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9DCE1")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for idx, (header, required, _values) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.fill = req_fill if required else opt_fill
        cell.font = head_font
        cell.alignment = head_align
        cell.border = border
        ws.column_dimensions[get_column_letter(idx)].width = WIDTHS.get(header, 18)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    last = MAX_ROWS + 1  # data rows 2..501
    for idx, (header, _required, values) in enumerate(COLUMNS, start=1):
        col = get_column_letter(idx)
        rng = f"{col}2:{col}{last}"
        if values:
            formula = '"' + ",".join(values) + '"'
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.errorStyle = "stop"
            dv.showErrorMessage = True
            dv.errorTitle = "Invalid value"
            dv.error = "Choose one of: " + ", ".join(values)
            dv.add(rng)
            ws.add_data_validation(dv)
        elif header.endswith("NO OF PIECES"):
            dv = DataValidation(type="whole", operator="greaterThanOrEqual",
                                formula1="1", allow_blank=True)
            dv.errorStyle = "stop"
            dv.showErrorMessage = True
            dv.errorTitle = "Invalid number"
            dv.error = "No. of pieces must be a whole number of at least 1."
            dv.add(rng)
            ws.add_data_validation(dv)
        elif header.endswith("WEIGHT (IN KG)"):
            dv = DataValidation(type="decimal", operator="greaterThan",
                                formula1="0", allow_blank=True)
            dv.errorStyle = "stop"
            dv.showErrorMessage = True
            dv.errorTitle = "Invalid weight"
            dv.error = "Weight (kg) must be a number greater than 0."
            dv.add(rng)
            ws.add_data_validation(dv)


def build_instructions_sheet(ws):
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 60

    title_font = Font(bold=True, size=16, color=PURPLE)
    sub_font = Font(size=10, color=GRAY)
    h2_font = Font(bold=True, size=12, color="FFFFFF")
    h2_fill = PatternFill("solid", fgColor=PURPLE)
    bullet_font = Font(size=10, color="222222")
    th_font = Font(bold=True, size=10, color="FFFFFF")
    th_fill = PatternFill("solid", fgColor=GRAY)
    req_font = Font(bold=True, size=10, color=ORANGE)
    opt_font = Font(size=10, color=GRAY)
    cell_font = Font(size=10, color="222222")
    wrap = Alignment(vertical="top", wrap_text=True)

    r = 1
    ws.cell(row=r, column=1, value="Fetcher Cargo — Bulk Shipment Upload").font = title_font
    r += 1
    ws.cell(row=r, column=1,
            value="Fill in the “Shipments” tab — one row per shipment — then upload it on the Bulk Create page.").font = sub_font
    r += 2

    def section(title):
        nonlocal r
        c = ws.cell(row=r, column=1, value=title)
        c.font = h2_font
        c.fill = h2_fill
        for col in (2, 3):
            ws.cell(row=r, column=col).fill = h2_fill
        r += 1

    def bullet(text):
        nonlocal r
        c = ws.cell(row=r, column=1, value="•  " + text)
        c.font = bullet_font
        c.alignment = wrap
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        r += 1

    section("Before you start")
    for b in [
        "Enter one shipment per row on the “Shipments” tab, starting at row 2.",
        "Do NOT rename, reorder, or delete the header row (row 1).",
        "Up to 500 shipments per file.",
        "Required columns have a PURPLE header; optional columns have a GRAY header.",
        "Drop-down columns only accept the listed values (pick from the arrow).",
        "A shipment can have up to 5 parcels: fill PARCEL 1, then add PARCEL 2–5 only if there are more boxes.",
        "Leave AWB, Client ID, Status, Billing, etc. out — they are assigned automatically by Fetcher / ops.",
        "REF is your own optional reference; we echo it back in the results so you can match any errors to your rows.",
    ]:
        bullet(b)
    r += 1

    section("Columns")
    # table header
    for col, label in [(1, "Column"), (2, "Required"), (3, "Allowed values / format")]:
        c = ws.cell(row=r, column=col, value=label)
        c.font = th_font
        c.fill = th_fill
        c.alignment = wrap
    r += 1
    for header, required, _values in COLUMNS:
        ws.cell(row=r, column=1, value=header).font = cell_font
        rc = ws.cell(row=r, column=2, value="Required" if required else "Optional")
        rc.font = req_font if required else opt_font
        nc = ws.cell(row=r, column=3, value=NOTES.get(header, ""))
        nc.font = cell_font
        nc.alignment = wrap
        r += 1
    r += 1

    section("Example row")
    example = (
        "REF=ORD-1001 | SCOPE=Domestic | PICK UP ADDRESS=12 MG Road, Bengaluru 560001 | "
        "PICK UP PINCODE/ZIP CODE=560001 | PICK UP CONTACT NO=9876543210 | "
        "PARCEL 1 NO OF PIECES=2 | PARCEL 1 WEIGHT (IN KG)=5.5 | PARCEL 1 DIMENSIONS (IN CM)=30x20x15 | "
        "PARCEL 2 NO OF PIECES=1 | PARCEL 2 WEIGHT (IN KG)=3 | "
        "DELIVERY ADDRESS=4 Park Street, Kolkata 700016 | "
        "DELIVERY PINCODE/ZIP CODE=700016 | DELIVERY CONTACT NO=9123456780 | "
        "SHIPMENT TYPE=Commercial | MODE=Express | SHIPMENT CATAGORY=Non-Doc | IS THIS A DG SHIPMENT?=No"
    )
    ec = ws.cell(row=r, column=1, value=example)
    ec.font = cell_font
    ec.alignment = wrap
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.row_dimensions[r].height = 60


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Shipments"
    build_shipments_sheet(ws)
    build_instructions_sheet(wb.create_sheet("Instructions"))
    wb.active = 0  # open on the Shipments tab

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    wb.save(OUT_PATH)
    print("wrote", OUT_PATH)


if __name__ == "__main__":
    main()
